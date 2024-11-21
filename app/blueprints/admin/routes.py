import io
import os
import tempfile
import zipfile
from datetime import datetime
from math import ceil
from threading import Thread

from flask import render_template, redirect, url_for, flash, request, Blueprint, current_app, jsonify
from flask_login import login_user, login_required, logout_user, current_user
from openpyxl.comments import Comment
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from werkzeug.security import check_password_hash, generate_password_hash

from app import db
from app.forms import AdminLoginForm, ChangePasswordForm, SurveyForm, CommitteeForm
from app.models import Admin, User, Survey, Committee, SurveyInstruction, Record, Feedback, SurveySettings
from app.schemas import SurveySchema, SurveyList
from app.utils import generate_random_reports, get_project_root

admin_bp = Blueprint('admin', __name__, template_folder='templates/admin')


@admin_bp.route('/')
@login_required
def index():
    # Общая статистика
    user_count = User.query.count()
    committee_count = Committee.query.count()
    survey_count = Survey.query.count()
    record_count = Record.query.count()

    # Получаем список файлов отчетов
    reports_folder = os.path.join(get_project_root(), current_app.config["REPORTS_FOLDER"])
    if not os.path.exists(reports_folder):
        os.makedirs(reports_folder)
    report_files = [f for f in os.listdir(reports_folder) if f.endswith('.xlsx') or f.endswith('.zip')]
    report_count = len(report_files)

    # Детали опросов
    surveys = Survey.query.all()
    survey_details = []
    for survey in surveys:
        directions = survey.directions  # Предполагается отношение один-к-многим
        direction_count = len(directions)
        question_count = 0
        for direction in directions:
            for criterion in direction.criterions:
                if criterion.subcriterions:
                    question_count += len(criterion.subcriterions)
                else:
                    question_count += 1  # Criterion без subcriterions считается как один вопрос
        survey_details.append({
            'title': survey.title,
            'direction_count': direction_count,
            'question_count': question_count,
        })

    # Детали комитетов
    committees = Committee.query.all()
    committee_details = []
    for committee in committees:
        user_count_in_committee = User.query.filter_by(committee_id=committee.id).count()
        committee_details.append({
            'name': committee.name,
            'user_count': user_count_in_committee,
        })

    return render_template('admin/dashboard.html',
                           title='Админка - Панель управления',
                           user_count=user_count,
                           committee_count=committee_count,
                           survey_count=survey_count,
                           report_count=report_count,
                           record_count=record_count,
                           surveys=survey_details,
                           committees=committee_details)


@admin_bp.route('/feedbacks')
@login_required
def view_feedbacks():
    feedbacks = Feedback.query.order_by(Feedback.submitted_at.desc()).all()
    return render_template('admin/feedbacks.html', feedbacks=feedbacks, title='Сообщения обратной связи')


@admin_bp.route('/login', methods=['GET', 'POST'])
def login():
    form = AdminLoginForm()
    if form.validate_on_submit():
        admin = Admin.query.filter_by(name=form.name.data).first()
        if admin and check_password_hash(admin.password, form.password.data):
            login_user(admin)
            flash('Вы успешно вошли в систему.', 'success')
            return redirect(url_for('admin.dashboard'))
        else:
            flash('Неверное имя пользователя или пароль.', 'danger')
    return render_template('admin/login.html', form=form)


@admin_bp.route('/logout')
@login_required
def logout():
    logout_user()
    flash('Вы вышли из системы.', 'success')
    return redirect(url_for('admin.login'))


@admin_bp.route('/dashboard')
@login_required
def dashboard():
    # Общая статистика
    user_count = User.query.count()
    committee_count = Committee.query.count()
    survey_count = Survey.query.count()
    record_count = Record.query.count()

    # Получаем список файлов отчетов
    reports_folder = os.path.join(get_project_root(), current_app.config["REPORTS_FOLDER"])
    if not os.path.exists(reports_folder):
        os.makedirs(reports_folder)
    report_files = [f for f in os.listdir(reports_folder) if f.endswith('.xlsx') or f.endswith('.zip')]
    report_count = len(report_files)

    # Детали опросов
    surveys = Survey.query.all()
    survey_details = []
    for survey in surveys:
        directions = survey.directions  # Предполагается отношение один-к-многим
        direction_count = len(directions)
        question_count = 0
        for direction in directions:
            for criterion in direction.criterions:
                if criterion.subcriterions:
                    question_count += len(criterion.subcriterions)
                else:
                    question_count += 1  # Criterion без subcriterions считается как один вопрос
        survey_details.append({
            'title': survey.title,
            'direction_count': direction_count,
            'question_count': question_count,
        })

    # Детали комитетов
    committees = Committee.query.all()
    committee_details = []
    for committee in committees:
        user_count_in_committee = User.query.filter_by(committee_id=committee.id).count()
        committee_details.append({
            'name': committee.name,
            'user_count': user_count_in_committee,
        })

    return render_template('admin/dashboard.html',
                           title='Админка - Панель управления',
                           user_count=user_count,
                           committee_count=committee_count,
                           survey_count=survey_count,
                           report_count=report_count,
                           record_count=record_count,
                           surveys=survey_details,
                           committees=committee_details)


@admin_bp.route('/change_password', methods=['GET', 'POST'])
@login_required
def change_password():
    form = ChangePasswordForm()
    if form.validate_on_submit():
        current_user.password = generate_password_hash(form.new_password.data, method='sha256')
        db.session.commit()
        flash('Ваш пароль был успешно обновлен!', 'success')
        return redirect(url_for('admin.dashboard'))
    return render_template('admin/change_password.html', form=form, title='Смена пароля')


@admin_bp.route('/survey', methods=['GET', 'POST'])
@login_required
def survey():
    form = SurveyForm()
    if form.validate_on_submit():
        new_survey = Survey(question=form.question.data, answer=form.answer.data)
        db.session.add(new_survey)
        db.session.commit()
        flash('Ответ был успешно записан.', 'success')
        return redirect(url_for('admin.survey'))
    return render_template('admin/survey.html', form=form, title='Опросник')


@admin_bp.route('/surveys', methods=['GET', 'POST'])
@login_required
def surveys():
    if request.method == "POST":
        form_data = request.form.to_dict()
        title = form_data.get("title")
        slug = form_data.get("slug")
        introduction = form_data.get("introduction", "")

        new_survey = Survey(title=title, slug=slug, introduction=introduction)
        db.session.add(new_survey)
        db.session.commit()

        flash("Новый опрос успешно создан", "success")
        return redirect(url_for('admin.surveys'))

    elif request.method == "GET":
        surveys_list = Survey.query.all()
        surveys = SurveyList.model_validate({'surveys': [survey.to_dict() for survey in surveys_list]})
        return render_template('admin/surveys.html', title='Опросники', surveys=surveys)


@admin_bp.route('/surveys/<int:survey_id>', methods=['GET'])
@login_required
def survey_by_id(survey_id: int):
    survey_list = Survey.query.get_or_404(survey_id)
    data = SurveySchema.model_validate(survey_list.to_dict())

    return render_template('admin/questions_tree.html', survey=data, title=data.title)


@admin_bp.route('/surveys/edit/<int:survey_id>', methods=['GET', 'POST'])
@login_required
def edit_survey(survey_id):
    survey = Survey.query.get_or_404(survey_id)

    if request.method == "POST":
        form_data = request.form.to_dict()
        survey.title = form_data.get("title", survey.title)
        survey.slug = form_data.get("slug", survey.slug)
        survey.introduction = form_data.get("introduction", survey.introduction)

        db.session.commit()
        flash("Настройки опроса успешно обновлены", "success")
        return redirect(url_for('admin.surveys'))

    return render_template('admin/edit_survey.html', survey=survey)


@admin_bp.route('/api/surveys/<int:survey_id>/settings/', methods=['PUT'])
@login_required
def update_survey_settings(survey_id):
    survey = Survey.query.get_or_404(survey_id)
    data = request.get_json()
    survey.slug = data.get('slug', survey.slug)
    settings_data = data.get('settings', {})
    for key, value in settings_data.items():
        setting = SurveySettings.query.filter_by(survey_id=survey_id, key=key).first()
        if setting:
            setting.value = value
    db.session.commit()
    return jsonify({'success': True})


# Обновление вступления
@admin_bp.route('/api/surveys/<int:survey_id>/introduction/', methods=['PUT'])
@login_required
def update_survey_introduction(survey_id):
    survey = Survey.query.get_or_404(survey_id)
    data = request.get_json()
    survey.introduction = data.get('introduction', survey.introduction)
    db.session.commit()
    return jsonify({'success': True})


# Обновление инструкции
@admin_bp.route('/api/instructions/<int:instruction_id>/', methods=['PUT'])
@login_required
def update_instruction(instruction_id):
    instruction = SurveyInstruction.query.get_or_404(instruction_id)
    data = request.get_json()
    instruction.title = data.get('title', instruction.title)
    instruction.body_html = data.get('body_html', instruction.body_html)
    db.session.commit()
    return jsonify({'success': True})


@admin_bp.route('/reports')
@login_required
def reports():
    page = request.args.get('page', 1, type=int)
    committee_filter = request.args.get('committee', type=int)
    sort_by = request.args.get('sort_by', 'id')  # Поле для сортировки, по умолчанию — ID
    sort_order = request.args.get('sort_order', 'desc')  # Порядок сортировки, по умолчанию — по убыванию

    # Создаем запрос с сортировкой и фильтрацией
    records_query = Record.query.join(User).join(Committee, User.committee_id == Committee.id)

    # Фильтр по комитету
    if committee_filter:
        records_query = records_query.filter(User.committee_id == committee_filter)

    # Сортировка по выбранному полю и порядку
    if sort_by == 'id':
        records_query = records_query.order_by(Record.id.desc() if sort_order == 'desc' else Record.id.asc())
    elif sort_by == 'committee':
        records_query = records_query.order_by(Committee.name.desc() if sort_order == 'desc' else Committee.name.asc())
    elif sort_by == 'created_at':  # Сортировка по дате создания
        records_query = records_query.order_by(
            Record.created_at.desc() if sort_order == 'desc' else Record.created_at.asc())

    records_paginated = records_query.paginate(page=page, per_page=10)

    committees = Committee.query.all()
    surveys = Survey.query.all()

    return render_template(
        'admin/reports.html',
        surveys=surveys,
        records=records_paginated.items,
        pagination=records_paginated,
        committees=committees,
        selected_committee=committee_filter,
        sort_by=sort_by,
        sort_order=sort_order,
    )


@admin_bp.route('/reports/delete/<int:record_id>', methods=['POST'])
@login_required
def delete_report(record_id):
    record = Record.query.get_or_404(record_id)
    db.session.delete(record)
    db.session.commit()
    flash("Отчет успешно удален", "success")
    return redirect(url_for('admin.reports'))


@admin_bp.route('/reports/view/<int:record_id>')
@login_required
def view_report(record_id):
    record = Record.query.get_or_404(record_id)
    survey = Survey.query.get_or_404(record.survey_id)
    return render_template('admin/view_report.html', record=record, survey=survey)


@admin_bp.route('/generate_reports')
@login_required
def generate_reports():
    try:
        generate_random_reports()
        flash('Отчёты успешно сгенерированы.', 'success')
    except Exception as e:
        flash(f'Ошибка при генерации отчётов: {str(e)}', 'error')
    return redirect(url_for('admin.dashboard'))


@admin_bp.route('/upload-methodology', methods=['GET', 'POST'])
@login_required
def upload_methodology():
    if request.method == 'POST':
        file = request.files['file']
        if file:
            file.save(f'/static/files/{file.filename}')
            flash('Файл методологии успешно загружен.', 'success')
        else:
            flash('Ошибка при загрузке файла.', 'danger')
    return render_template('admin/upload_methodology.html', title='Загрузка методологии')


@admin_bp.route('/download/<filetype>')
@login_required
def download(filetype):
    if filetype == 'xlsx':
        return send_file('/path/to/file.xlsx', as_attachment=True)
    elif filetype == 'json':
        return send_file('/path/to/file.json', as_attachment=True)
    elif filetype == 'database':
        return send_file('/path/to/database.db', as_attachment=True)
    else:
        flash('Неверный тип файла для загрузки.', 'danger')
        return redirect(url_for('admin.user_data'))


@admin_bp.route('/committees', methods=['GET'])
@login_required
def view_committees():
    form = CommitteeForm()
    form.parent_id.choices = [(0, 'Нет родительского комитета')] + [(committee.id, committee.name) for committee in
                                                                    Committee.query.all()]
    return render_template('admin/committees.html', form=form)


@admin_bp.route('/surveys/<int:survey_id>/edit_intro', methods=['GET', 'POST'])
@login_required
def edit_survey_intro(survey_id):
    survey = Survey.query.get_or_404(survey_id)
    if request.method == 'POST':
        survey.introduction = request.form.get('introduction')
        db.session.commit()
        flash('Вступление обновлено', 'success')
        return redirect(url_for('admin.survey_by_id', survey_id=survey_id))
    return render_template('admin/edit_intro.html', survey=survey)


@admin_bp.route('/surveys/<int:survey_id>/add_intro', methods=['GET', 'POST'])
@login_required
def add_survey_intro(survey_id):
    survey = Survey.query.get_or_404(survey_id)
    if request.method == 'POST':
        survey.introduction = request.form.get('introduction')
        db.session.commit()
        flash('Вступление добавлено', 'success')
        return redirect(url_for('admin.survey_by_id', survey_id=survey_id))
    return render_template('admin/add_intro.html', survey=survey)


@admin_bp.route('/instructions/<int:instruction_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_instruction(instruction_id):
    instruction = SurveyInstruction.query.get_or_404(instruction_id)
    if request.method == 'POST':
        instruction.title = request.form.get('title')
        instruction.body_html = request.form.get('body_html')
        db.session.commit()
        flash('Инструкция обновлена', 'success')
        return redirect(url_for('admin.survey_by_id', survey_id=instruction.survey_id))
    return render_template('admin/edit_instruction.html', instruction=instruction)


@admin_bp.route('/surveys/<int:survey_id>/add_instruction', methods=['GET', 'POST'])
@login_required
def add_instruction(survey_id):
    if request.method == 'POST':
        new_instruction = SurveyInstruction(
            title=request.form.get('title'),
            body_html=request.form.get('body_html'),
            survey_id=survey_id
        )
        db.session.add(new_instruction)
        db.session.commit()
        flash('Инструкция добавлена', 'success')
        return redirect(url_for('admin.survey_by_id', survey_id=survey_id))
    return render_template('admin/add_instruction.html', survey_id=survey_id)


def load_survey_data(survey_id):
    survey_model = Survey.query.get(survey_id)
    if not survey_model:
        raise ValueError("Survey not found")
    return SurveySchema.model_validate(survey_model.to_dict())


def prepare_directions_structure(survey):
    directions = {}
    for direction in survey.directions:
        directions[direction.id] = {
            "direction_id": direction.id,
            "direction_title": direction.title,
            "coefficient": direction.coefficient,
            "criteria": [],
            "user_totals": {}
        }
        for criterion in direction.criterions:
            criterion_data = {
                "number": criterion.number,
                "weight": criterion.weight,
                "criterion_title": criterion.title,
                "subcriteria": [],
                "answers": []
            }
            for subcriterion in criterion.subcriterions:
                subcriterion_data = {
                    "subcriterion_title": subcriterion.title,
                    "subcriterion_weight": subcriterion.weight,
                    "answers": []
                }
                criterion_data["subcriteria"].append(subcriterion_data)
            directions[direction.id]["criteria"].append(criterion_data)
    return directions


def fill_criteria_with_answers(records, directions):
    for record in records:
        user_id = record.user.id
        user_info = {
            'id': user_id,
            'post': record.user.post,
            'subdivision': record.user.subdivision
        }

        # Инициализируем словарь для пользователя в каждом направлении
        for direction_id in directions.keys():
            if user_id not in directions[direction_id]['user_totals']:
                directions[direction_id]['user_totals'][user_id] = {
                    'total_value': 0,
                    'user_info': user_info
                }

        # Обрабатываем ответы пользователя
        for answer in record.answers:
            criterion = answer.criterion
            subcriterion = answer.subcriterion
            total_value = (answer.answer_value or 0) + (answer.range_value or 0)
            weight = subcriterion.weight if subcriterion else criterion.weight if criterion else 0
            weighted_value = total_value * weight
            direction_id = criterion.direction_id if criterion else None

            if direction_id and direction_id in directions:
                # Добавляем к общей сумме пользователя для сводного листа
                directions[direction_id]['user_totals'][user_id]['total_value'] += weighted_value

                # Добавляем ответы в критерии и подкритерии для детального листа
                for criterion_data in directions[direction_id]["criteria"]:
                    if criterion.number == criterion_data["number"]:
                        if subcriterion:
                            for subcriterion_data in criterion_data["subcriteria"]:
                                if subcriterion.title == subcriterion_data["subcriterion_title"]:
                                    subcriterion_data["answers"].append({
                                        "user_id": user_id,
                                        "answer_value": total_value,
                                        "comment": answer.comment
                                    })
                        else:
                            criterion_data["answers"].append({
                                "user_id": user_id,
                                "answer_value": total_value,
                                "comment": answer.comment
                            })


def generate_summary_sheet(wb, directions, users):
    ws = wb.create_sheet("Свод по разделам")

    row_index = 2
    for direction_data in directions.values():
        # Объединяем ячейки для названия направления
        ws.merge_cells(start_row=row_index, start_column=3, end_row=row_index, end_column=6)
        ws[f"C{row_index}"] = direction_data["direction_title"]
        ws[f"C{row_index}"].font = Font(bold=True)
        row_index += 1

        # Устанавливаем заголовки столбцов
        ws[f"C{row_index}"] = "Наименование направления"
        ws[f"E{row_index}"] = "Подразделение"
        ws[f"F{row_index}"] = "Значение показателя"
        for col in ['C', 'E', 'F']:
            ws[f"{col}{row_index}"].font = Font(bold=True)
            ws[f"{col}{row_index}"].alignment = Alignment(horizontal='center')
        row_index += 1

        # Получаем и сортируем данные пользователей
        user_totals = list(direction_data.get('user_totals', {}).values())
        user_totals.sort(key=lambda x: x['total_value'], reverse=True)

        start_data_row = row_index
        for user_data in user_totals:
            ws[
                f"E{row_index}"] = f"{user_data['user_info']['subdivision']}_{user_data['user_info']['post']}_{user_data['user_info']['id']}"
            ws[f"F{row_index}"] = round(user_data['total_value'], 2)
            row_index += 1

        # Строка "Итого по выборке"
        ws[f"C{row_index}"] = "Итого по выборке"
        row_index += 1

        # Статистические расчёты
        data_range = f"F{start_data_row}:F{row_index - 2}"
        ws[f"C{row_index}"] = "Минимальное значение"
        ws[f"F{row_index}"] = f"=MIN({data_range})"
        row_index += 1

        ws[f"C{row_index}"] = "Среднее значение"
        ws[f"F{row_index}"] = f"=AVERAGE({data_range})"
        row_index += 1

        ws[f"C{row_index}"] = "Максимальное значение"
        ws[f"F{row_index}"] = f"=MAX({data_range})"
        row_index += 1

        ws[f"C{row_index}"] = "1-й квартиль (25-й процентиль)"
        ws[f"F{row_index}"] = f"=ПРОЦЕНТИЛЬ.ВКЛ({data_range},0.25)"
        row_index += 1

        ws[f"C{row_index}"] = "2-й квартиль (медиана)"
        ws[f"F{row_index}"] = f"=MEDIAN({data_range})"
        row_index += 1

        ws[f"C{row_index}"] = "3-й квартиль  (75-й процентиль)"
        ws[f"F{row_index}"] = f"=ПРОЦЕНТИЛЬ.ВКЛ({data_range},0.75)"
        row_index += 1

        ws[f"C{row_index}"] = "Среднеквадратичное (стандартное) отклонение"
        ws[f"F{row_index}"] = f"=STDEV({data_range})"
        row_index += 1

        ws[f"C{row_index}"] = "Коэффициент вариации"
        ws[f"F{row_index}"] = f"=F{row_index - 1}/F{row_index - 7}"
        row_index += 1

        ws[f"C{row_index}"] = "Оценка коэффициента вариации"
        ws[f"F{row_index}"] = (
            f'=IF(F{row_index - 1}>0.33, "выборка неоднородна", '
            f'IF(F{row_index - 1}>0.2, "однородность выборки низкая", '
            f'IF(F{row_index - 1}>0.1, "однородность выборки средняя", "однородность выборки высокая")))'
        )
        row_index += 2  # Пропускаем строку перед следующим направлением


def generate_excel_report(wb, directions, users):
    ws = wb.active
    ws.title = "Детализация по разделам"

    headers = [
        "Наименование направления", "Показатель ИОГВ, баллы ИТОГ", "Вес ИТОГ для главного свода",
        "Вес критерий", "№", "Критерий", "Наименование подкритерия", "Значение показателя", ""
    ]

    user_headers = [f"{user['subdivision']} {user['post']} ({user['id']})" for user in users]
    user_ids = [user['id'] for user in users]

    headers.extend(user_headers)

    ws.append(headers)

    # Установка стиля для заголовков
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num).font = Font(bold=True)
        ws.cell(row=1, column=col_num).alignment = Alignment(horizontal='center', wrap_text=True)

    row_index = 2
    for direction_data in directions.values():
        direction_start_row = row_index

        # Подсчитываем общее количество строк для направления
        num_rows_for_direction = sum(
            len(criterion_data["subcriteria"]) if criterion_data["subcriteria"] else 1
            for criterion_data in direction_data["criteria"]
        )

        # Запись информации по направлению
        ws[f"A{row_index}"] = f"{direction_data['direction_id']} {direction_data['direction_title']}"

        # Устанавливаем коэффициент направления (например, 0.15 для 15%)
        coefficient = direction_data['coefficient']
        if coefficient > 1:
            # Если коэффициент указан в процентах (например, 15 для 15%), делим на 100
            coefficient /= 100

        ws[f"C{row_index}"] = coefficient
        ws[f"C{row_index}"].number_format = '0%'

        ws[f"A{row_index}"].alignment = Alignment(horizontal='center', vertical='center')
        ws[f"C{row_index}"].alignment = Alignment(horizontal='center', vertical='center')

        # Объединение ячеек для направления
        ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index + num_rows_for_direction, end_column=1)
        ws.merge_cells(start_row=row_index, start_column=2, end_row=row_index + num_rows_for_direction, end_column=2)
        ws.merge_cells(start_row=row_index, start_column=3, end_row=row_index + num_rows_for_direction, end_column=3)

        # Сохраняем строку с названием направления для итоговых вычислений
        direction_title_row = row_index

        for criterion_data in direction_data["criteria"]:
            subcriteria = criterion_data["subcriteria"]
            criterion_weight = criterion_data["weight"]
            criterion_number = criterion_data["number"]
            criterion_title = criterion_data["criterion_title"]
            if subcriteria:
                num_subcriteria = len(subcriteria)
                first_subcriterion_row = row_index

                for idx, subcriterion_data in enumerate(subcriteria):
                    subcriterion_weight = subcriterion_data["subcriterion_weight"]
                    subcriterion_title = subcriterion_data["subcriterion_title"]
                    ws[f"D{row_index}"] = subcriterion_weight

                    # Заполняем номер и название критерия только в первую строку группы
                    # if idx == 0:
                    ws[f"E{row_index}"] = criterion_number
                    ws[f"F{row_index}"] = criterion_title
                    # Записываем вес критерия только в первую строку группы
                    # ws[f"D{row_index}"] = criterion_weight
                    # else:
                    #     ws[f"E{row_index}"] = ''
                    #     ws[f"F{row_index}"] = ''
                    # Оставляем ячейку веса критерия пустой
                    # ws[f"D{row_index}"] = ''

                    ws[f"G{row_index}"] = subcriterion_title

                    # Заполнение ответов пользователей
                    fill_user_responses_detail(ws, row_index, user_ids, subcriterion_data["answers"])

                    # Расчет среднего значения для подкритерия
                    first_user_col = 10  # Колонка 'J'
                    last_user_col = 9 + len(user_ids)
                    ws[
                        f"H{row_index}"] = f"=AVERAGE({get_column_letter(first_user_col)}{row_index}:{get_column_letter(last_user_col)}{row_index})"

                    row_index += 1

                # Объединение ячеек для критериев
                ws.merge_cells(start_row=first_subcriterion_row, start_column=5, end_row=row_index - 1,
                               end_column=5)  # Номер
                ws.merge_cells(start_row=first_subcriterion_row, start_column=6, end_row=row_index - 1,
                               end_column=6)  # Критерий

                # Вес критерия записан в первую строку группы
            else:
                # Если нет подкритериев
                ws[f"D{row_index}"] = criterion_weight
                ws[f"E{row_index}"] = criterion_number
                ws[f"F{row_index}"] = criterion_title
                ws[f"G{row_index}"] = ''

                # Заполнение ответов пользователей
                fill_user_responses_detail(ws, row_index, user_ids, criterion_data["answers"])

                # Расчет среднего значения для критерия
                first_user_col = 10  # Колонка 'J'
                last_user_col = 9 + len(user_ids)
                ws[
                    f"H{row_index}"] = f"=AVERAGE({get_column_letter(first_user_col)}{row_index}:{get_column_letter(last_user_col)}{row_index})"

                row_index += 1

        # Итог по направлению для пользователей
        total_row = row_index
        ws[f"F{row_index}"] = "Итого"
        ws[f"F{row_index}"].font = Font(bold=True)
        ws[f"F{row_index}"].alignment = Alignment(horizontal='center')

        # Расчет среднего значения показателя направления
        # Среднее значение из колонки "Значение показателя" для всех строк направления
        value_cells = [f"H{r}" for r in range(direction_start_row, row_index)]
        ws[f"H{row_index}"] = f"=AVERAGE({','.join(value_cells)})"
        ws[f"H{row_index}"].number_format = '0.00'

        # Рассчитываем взвешенные суммы для каждого пользователя
        calculate_weighted_sums(ws, direction_start_row, row_index - 1, len(user_ids))

        # Рассчитываем показатель ИОГВ, баллы ИТОГ для направления
        first_user_col = 10  # Колонка 'J'
        last_user_col = 9 + len(user_ids)
        user_total_cells = [f"{get_column_letter(col)}{total_row}" for col in range(first_user_col, last_user_col + 1)]
        ws[f"B{direction_title_row}"] = f"=AVERAGE({','.join(user_total_cells)})"
        ws[f"B{direction_title_row}"].number_format = '0.00'

        row_index += 1  # Переходим к следующей строке после "Итого"

    # Добавляем финальный итог
    add_final_total(ws, row_index)


def fill_direction_data(ws, row_index, direction_data, user_headers):
    start_row = row_index
    num_rows_for_direction = sum(
        len(criterion_data["subcriteria"]) if criterion_data["subcriteria"] else 1
        for criterion_data in direction_data["criteria"]
    )

    # Запись информации по направлению
    ws[f"A{row_index}"] = f"{direction_data['direction_id']} {direction_data['direction_title']}"
    ws[
        f"B{row_index}"] = f"=AVERAGE(J{row_index + num_rows_for_direction}:CM{row_index + num_rows_for_direction})"  # Среднее для итоговой строки критерия
    ws[f"C{row_index}"] = f"{direction_data['coefficient']}%"
    ws[f"A{row_index}"].alignment = Alignment(horizontal='center')
    ws[f"C{row_index}"].number_format = '0%'  # Форматирование как проценты

    for criterion_data in direction_data["criteria"]:
        subcriteria = criterion_data["subcriteria"]

        # Если у критерия нет подкритериев, вес идет в столбец D для самого критерия
        ws[f"D{row_index}"] = criterion_data["weight"]
        ws[f"E{row_index}"] = criterion_data["number"]
        ws[f"F{row_index}"] = criterion_data["criterion_title"]

        if subcriteria:
            for subcriterion_data in subcriteria:
                ws[f"G{row_index}"] = subcriterion_data["subcriterion_title"]
                ws[f"D{row_index}"] = subcriterion_data["subcriterion_weight"]

                # Формула с диапазоном для расчета среднего
                answer_start_col = 10  # J колонка
                answer_end_col = answer_start_col + len(user_headers) - 1
                ws[f"H{row_index}"] = f"=AVERAGE(J{row_index}:CM{row_index})"  # AVERAGE диапазон для среднего

                fill_user_responses(ws, row_index, user_headers, criterion_data["answers"],
                                    subcriterion_data["subcriterion_title"], subcriterion_data["subcriterion_weight"])
                row_index += 1
        else:
            answer_start_col = 10
            answer_end_col = answer_start_col + len(user_headers) - 1
            ws[f"H{row_index}"] = f"=AVERAGE(J{row_index}:CM{row_index})"

            fill_user_responses(ws, row_index, user_headers, criterion_data["answers"],
                                criterion_data["criterion_title"], criterion_data["weight"])
            row_index += 1

    calculate_weighted_sums(ws, start_row, row_index, user_headers)

    # Итог по направлению
    ws[f"F{row_index}"] = "Итого"
    ws[f"H{row_index}"] = f"=AVERAGE(H{start_row}:H{row_index - 1})"
    ws[f"F{row_index}"].font = Font(bold=True)
    ws[f"F{row_index}"].alignment = Alignment(horizontal='center')
    row_index += 1

    return row_index


def fill_user_responses_detail(ws, row_index, user_ids, answers):
    user_answers = {ans['user_id']: ans for ans in answers}
    for idx, user_id in enumerate(user_ids):
        col = 10 + idx  # Начинаем с колонки 'J' (10)
        answer_data = user_answers.get(user_id, {"answer_value": 0, "comment": None})
        user_answer = answer_data["answer_value"]

        # Вставляем значение в ячейку и форматируем
        cell = ws.cell(row=row_index, column=col)
        cell.value = user_answer
        cell.number_format = '0.00'

        # Добавляем комментарий, если он есть
        if answer_data["comment"]:
            comment_text = answer_data["comment"]
            cell.comment = Comment(comment_text, "User")


def fill_user_responses(ws, row_index, user_headers, answers, subcriterion_title, weight):
    user_answers = {
        f"{ans['user_subdivision']} {ans['user_post']} ({ans['user_id']})": {
            "value": ans["answer_value"],
            "comment": ans.get("comment")  # Извлекаем комментарий, если он есть
        }
        for ans in answers if ans["subcriterion_title"] == subcriterion_title
    }

    for col, header in enumerate(user_headers, start=10):
        answer_data = user_answers.get(header,
                                       {"value": 0, "comment": None})  # Если нет данных, значение 0 и без комментария
        user_answer = answer_data["value"]

        # Вставляем значение в ячейку и форматируем
        cell = ws.cell(row=row_index, column=col)
        cell.value = user_answer
        cell.number_format = '0.00'

        # Добавляем комментарий, если он есть
        if answer_data["comment"]:
            comment_text = answer_data["comment"]
            cell.comment = Comment(comment_text, "User")


def calculate_weighted_sums(ws, start_row, end_row, num_users):
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    for idx in range(num_users):
        col = 10 + idx  # Начинаем с колонки 'J' (10)
        user_total_cell = ws.cell(row=end_row + 1, column=col)
        user_total_cell.fill = yellow_fill
        user_total_cell.number_format = '0.00'

        # Формула для расчёта взвешенной суммы для пользователя
        formula_parts = []
        for row in range(start_row, end_row + 1):
            weight_cell = ws.cell(row=row, column=4)  # Колонка 'D' (Вес критерий)
            user_answer_cell = ws.cell(row=row, column=col)
            if weight_cell.value not in [None, '', 0]:
                formula_parts.append(f"{user_answer_cell.coordinate}*{weight_cell.coordinate}")
        formula = "+".join(formula_parts)
        user_total_cell.value = f"=({formula})" if formula else 0


def add_final_total(ws, row_index):
    total_cell = ws.cell(row=row_index, column=1)
    total_cell.value = "ИОГВ итог"
    total_cell.font = Font(bold=True)
    total_cell.alignment = Alignment(horizontal='center')

    # Рассчитываем общий итог
    direction_rows = []
    current_row = 2
    while current_row < row_index:
        direction_title = ws.cell(row=current_row, column=1).value
        if direction_title:
            direction_rows.append(current_row)
        # Пропускаем к следующему направлению
        current_row += 1
        while current_row < row_index and not ws.cell(row=current_row, column=1).value:
            current_row += 1

    # Формула для расчёта общего итога
    total_formula_parts = []
    for direction_row in direction_rows:
        coefficient_cell = ws.cell(row=direction_row, column=3)  # Колонка 'C' (Вес ИТОГ для главного свода)
        direction_value_cell = ws.cell(row=direction_row, column=2)  # Колонка 'B' (Показатель ИОГВ, баллы ИТОГ)
        total_formula_parts.append(f"{coefficient_cell.coordinate}*{direction_value_cell.coordinate}")

    total_formula = "+".join(total_formula_parts)
    total_value_cell = ws.cell(row=row_index, column=2)
    total_value_cell.value = f"=({total_formula})"
    total_value_cell.number_format = '0.00'


def calculate_weighted_sum(answers):
    """
    Рассчитывает сумму произведений веса и ответа для критериев и подкритериев в ответах пользователя.
    """
    total = 0
    for answer in answers:
        total += answer["answer_value"] * answer["weight"]
    return total


def generate_complete_report(filepath, directions, users):
    wb = Workbook()
    generate_excel_report(wb, directions, users)
    generate_summary_sheet(wb, directions, users)

    with io.BytesIO() as file_stream:
        wb.save(file_stream)
        file_stream.seek(0)
        with open(filepath, 'wb') as f:
            f.write(file_stream.read())


def generate_report_task(app, survey_id, committee_id):
    with app.app_context():
        try:
            reports_folder = os.path.join(get_project_root(), app.config["REPORTS_FOLDER"])
            if not os.path.exists(reports_folder):
                os.makedirs(reports_folder)

            timestamp = datetime.now().strftime("%Y%m%d%H%M%S")

            if committee_id == 'all':
                # Создаем временную директорию для хранения отчётов
                with tempfile.TemporaryDirectory() as temp_dir:
                    print(f"Genarete temp dir: {temp_dir}")

                    committees = Committee.query.all()
                    for committee in committees:
                        filename = f"{survey_id}-{committee.name}-{timestamp}.xlsx"
                        filepath = os.path.join(temp_dir, filename)
                        generate_single_report(filepath, survey_id, committee.id)

                    zip_filename = f"all_committees-{survey_id}-{timestamp}.zip"
                    zip_filepath = os.path.join(reports_folder, zip_filename)
                    with zipfile.ZipFile(zip_filepath, 'w', zipfile.ZIP_DEFLATED) as zipf:
                        for root, dirs, files in os.walk(temp_dir):
                            for file in files:
                                file_path = os.path.join(root, file)
                                arcname = os.path.basename(file_path)  # Имя файла внутри архива
                                zipf.write(file_path, arcname)
                # После выхода из блока with временная директория автоматически удалится
                return zip_filepath
            else:
                committee = Committee.query.get(committee_id)
                filename = f"{committee.name}-{survey_id}-{timestamp}.xlsx"
                filepath = os.path.join(reports_folder, filename)
                generate_single_report(filepath, survey_id, committee_id)
                return filepath
        except Exception as e:
            print(f"Error generating report: {e}")
            raise


def generate_single_report(filepath, survey_id, committee_id):
    try:
        print(f"Genarete filenane: {filepath}")
        survey = load_survey_data(survey_id)
        committee = Committee.query.get(committee_id)

        if not committee:
            return

        records_query = Record.query.join(User).filter(Record.survey_id == survey_id)
        if committee:
            records_query = records_query.filter(User.committee_id == committee.id)
        records = records_query.all()

        directions = prepare_directions_structure(survey)
        fill_criteria_with_answers(records, directions)

        users = [{"id": record.user.id, "post": record.user.post, "subdivision": record.user.subdivision} for record in
                 records]

        generate_complete_report(filepath, directions, users)

    except Exception as e:
        print(f"Error in generate_single_report: {e}")
        raise


@admin_bp.route('/generate_report', methods=['POST'])
@login_required
def generate_report():
    survey_id = request.form['survey_id']
    committee_id = request.form['committee_id']

    thread = Thread(target=generate_report_task, args=(current_app._get_current_object(), survey_id, committee_id))
    thread.start()

    flash('Отчет генерируется. Вы сможете скачать его через несколько минут.')
    return redirect(url_for('admin.reports'))


@admin_bp.route('/view_generated_reports')
@login_required
def view_generated_reports():
    page = request.args.get('page', 1, type=int)
    per_page = 10

    reports_folder = os.path.join(get_project_root(), current_app.config["REPORTS_FOLDER"])

    if not os.path.exists(reports_folder):
        os.makedirs(reports_folder)

    all_files = []
    for filename in os.listdir(reports_folder):
        if filename.endswith(".xlsx") or filename.endswith(".zip"):
            try:
                parts = filename.rsplit('-', 2)
                if len(parts) == 3:
                    committee_name, survey_id, timestamp_str = parts
                    survey_id = survey_id if survey_id.isdigit() else "Неизвестно"
                    timestamp = datetime.strptime(timestamp_str.replace(".xlsx", "").replace(".zip", ""),
                                                  "%Y%m%d%H%M%S")
                else:
                    committee_name, survey_id, timestamp = "Неизвестно", "Неизвестно", "Неизвестно"
            except (IndexError, ValueError):
                committee_name, survey_id, timestamp = "Неизвестно", "Неизвестно", "Неизвестно"

            survey = Survey.query.get(survey_id)
            all_files.append({
                "filename": filename,
                "committee_name": committee_name if committee_name != 'all_committees' else 'Все комитеты',
                "survey": survey.title if survey else "Неизвестно",
                "timestamp": timestamp if isinstance(timestamp, datetime) else "Неизвестно",
            })

    all_files.sort(key=lambda x: x["timestamp"] if isinstance(x["timestamp"], datetime) else datetime.min, reverse=True)

    total_files = len(all_files)
    total_pages = ceil(total_files / per_page)
    files_paginated = all_files[(page - 1) * per_page: page * per_page]

    return render_template(
        'admin/generated_reports.html',
        files=files_paginated,
        page=page,
        total_pages=total_pages,
    )


@admin_bp.route('/delete_generated_report/<filename>', methods=['POST'])
@login_required
def delete_generated_report(filename):
    reports_folder = os.path.join(get_project_root(), current_app.config["REPORTS_FOLDER"])
    filepath = os.path.join(reports_folder, filename)

    if os.path.exists(filepath):
        os.remove(filepath)
        flash(f'Файл {filename} успешно удален.', 'success')
    else:
        flash(f'Файл {filename} не найден.', 'error')

    return redirect(url_for('admin.view_generated_reports'))


@admin_bp.route('/download_report/<filename>')
@login_required
def download_report(filename):
    reports_folder = os.path.join(get_project_root(), current_app.config["REPORTS_FOLDER"])
    filepath = os.path.join(reports_folder, filename)
    return send_file(filepath, as_attachment=True)


@admin_bp.route('/delete_all_reports', methods=['POST'])
@login_required
def delete_all_reports():
    try:
        all_records = Record.query.all()

        for record in all_records:
            user = record.user
            db.session.delete(record)
            if user:
                db.session.delete(user)

        db.session.commit()

        flash('Все отчеты и связанные пользователи успешно удалены.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Произошла ошибка при удалении отчетов: {str(e)}', 'error')

    return redirect(url_for('admin.reports'))


from flask import send_file


@admin_bp.route('/download_database')
@login_required
def download_database():
    db_path = os.path.join(get_project_root(), current_app.config['SQLALCHEMY_DATABASE_URI'].replace('sqlite:///', ''))
    return send_file(db_path, as_attachment=True, attachment_filename='database.sqlite')
